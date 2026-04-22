import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from datetime import datetime
import os


def export_to_excel():
    patent_numbers = text_input.get("1.0", tk.END).strip().split("\n")

    # Remove blank lines
    patent_numbers = [p.strip() for p in patent_numbers if p.strip()]

    if not patent_numbers:
        messagebox.showwarning("No Input", "Please enter at least one patent or publication number.")
        return

    wb = Workbook()
    ws = wb.active

    # Set headers
    headers = ["Ref No.", "PDF Link", "Google Link", "Espacenet Link", "USPTO Link", "ABSTRACT"]
    column_widths = [25, 40, 40, 40, 40, 70]

    ws.append(headers)

    for col_num, width in enumerate(column_widths, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width

    # Style header
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Fill rows
    for number in patent_numbers:
        cleaned_number = number.strip()
        if not cleaned_number:
            continue

        # Remove "US" prefix to get the patent number
        uspto_number = cleaned_number.replace("US", "")
        
        # Build URLs
        pdf_url = f"https://image-ppubs.uspto.gov/dirsearch-public/print/downloadPdf/{uspto_number}"
        google_url = f"https://patents.google.com/patent/{cleaned_number}"
        espacenet_url = f"https://worldwide.espacenet.com/patent/search?q={cleaned_number}"
        uspto_url = f"https://ppubs.uspto.gov/pubwebapp/external.html?q={uspto_number}.pn."

        # Create row
        row = [cleaned_number, uspto_number, cleaned_number, cleaned_number, cleaned_number, ""]

        ws.append(row)
        row_idx = ws.max_row

        # Set hyperlinks
        ws.cell(row=row_idx, column=2).hyperlink = pdf_url
        ws.cell(row=row_idx, column=2).style = "Hyperlink"

        ws.cell(row=row_idx, column=3).hyperlink = google_url
        ws.cell(row=row_idx, column=3).style = "Hyperlink"

        ws.cell(row=row_idx, column=4).hyperlink = espacenet_url
        ws.cell(row=row_idx, column=4).style = "Hyperlink"

        ws.cell(row=row_idx, column=5).hyperlink = uspto_url
        ws.cell(row=row_idx, column=5).style = "Hyperlink"

    # Timestamped filename
    now = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"patent_export_{now}.xlsx"

    # Path to Downloads
    downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
    full_path = os.path.join(downloads_path, filename)

    # Save the workbook
    try:
        wb.save(full_path)
        messagebox.showinfo("Export Successful", f"Data exported to:\n{full_path}")
    except Exception as e:
        messagebox.showerror("Error", f"Could not save file:\n{e}")


# Tkinter UI
root = tk.Tk()
root.title("Patent Exporter")

tk.Label(root, text="Enter patent/publication numbers (one per line):").pack(pady=5)
text_input = tk.Text(root, height=15, width=45)
text_input.pack(padx=10)
tk.Button(root, text="Export to Excel", command=export_to_excel).pack(pady=10)

root.mainloop()